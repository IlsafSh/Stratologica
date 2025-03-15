# Функции загрузки и сохранения файлов

import tkinter as tk
from tkinter import filedialog, messagebox
import csv
import openpyxl

def load_matrix_from_file(root):
    """Открывает диалог выбора файла и загружает матрицу из .txt или .xlsx"""
    file_path = filedialog.askopenfilename(
        title="Выберите файл",
        filetypes=[("Текстовые файлы", "*.txt"), ("Excel файлы", "*.xlsx")]
    )

    if not file_path:
        return None  # Если файл не выбран, ничего не делать

    try:
        if file_path.endswith(".txt"):
            with open(file_path, "r", encoding="utf-8") as file:
                matrix = []
                for line in file:
                    # Разбиваем строку по табуляции и преобразуем в числа
                    row = [float(x.strip()) for x in line.strip().split('\t') if x.strip()]
                    if row:  # Добавляем только непустые строки
                        matrix.append(row)

        elif file_path.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            matrix = [[cell.value for cell in row] for row in sheet.iter_rows()]
            matrix = [[0 if value is None else float(value) for value in row] for row in matrix]  # Заменяем None на 0

        # Проверяем, что все строки имеют одинаковую длину
        if matrix and not all(len(row) == len(matrix[0]) for row in matrix):
            raise ValueError("Некорректный формат матрицы: строки разной длины")

        messagebox.showinfo("Успех", "Файл успешно загружен.")
        return matrix

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить файл: {e}")
        return None


def save_matrix_to_file(matrix, root):
    """Сохраняет текущую матрицу в .txt или .xlsx"""
    file_path = filedialog.asksaveasfilename(
        title="Сохранить файл как",
        defaultextension=".txt",
        filetypes=[("Текстовые файлы", "*.txt"), ("Excel файлы", "*.xlsx")]
    )

    if not file_path:
        return  # Если файл не выбран, ничего не делать

    try:
        if file_path.endswith(".txt"):
            with open(file_path, "w", encoding="utf-8") as file:
                for row in matrix:
                    # Используем табуляцию как разделитель
                    file.write('\t'.join(str(x) for x in row) + '\n')

        elif file_path.endswith(".xlsx"):
            wb = openpyxl.Workbook()
            sheet = wb.active
            for row in matrix:
                sheet.append(row)
            wb.save(file_path)

        messagebox.showinfo("Успех", "Файл успешно сохранён.")

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")
